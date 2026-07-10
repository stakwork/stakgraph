#!/usr/bin/env python3
"""
build_workbook.py — openpyxl workbook builder for the generate_xlsx agent tool.

Usage: python3 build_workbook.py <input.json>

Input JSON schema:
{
  "output": "/path/to/output.xlsx",
  "filename": "optional-hint",   // ignored; output path is canonical
  "sheets": [
    {
      "name": "Sheet1",
      "rows": [["Header A", "Header B"], [1, 2], ...],  // optional row-based data
      "cells": [                                          // optional fine-grained cells
        { "ref": "A1", "value": "Label" },
        { "ref": "B2", "formula": "=Sheet2!B2+1" }
      ],
      "computed": [                                       // optional computed cells (literal values only)
        {
          "ref": "B10",
          "op": "sum",            // "sum" | "percent_of_total" | "ratio"
          "range": "B2:B9"        // for sum (column or row)
        },
        {
          "ref": "C2",
          "op": "percent_of_total",
          "value_ref": "B2",      // numerator cell
          "total_ref": "B10",     // denominator cell
          "decimals": 2,          // default 2
          "as_fraction": false    // default false => percent-scaled (x100)
        },
        {
          "ref": "D2",
          "op": "ratio",
          "value_ref": "B2",
          "denominator_ref": "C2",
          "decimals": 4
        }
      ]
    }
  ]
}

- "rows" are written first (1-indexed, starting at row 1).
- "cells" are applied after rows and may override individual cells.
- "formula" values are written as-is (including cross-sheet refs like =Sheet2!B2).
- Both "value" and "formula" may coexist in a cell entry; "formula" takes precedence.
- "computed" entries are evaluated in a second pass (after all sheets/cells are written),
  sequentially in list order so earlier results feed later computations.
- Sheet-qualified refs are supported: "Sheet2!B2" or "Sheet2!B2:B9".
"""
import sys
import json
import os
import re

try:
    from openpyxl import Workbook
except ImportError as e:
    print(f"openpyxl is not installed: {e}", file=sys.stderr)
    sys.exit(1)

# A1 reference patterns
_CELL_RE = re.compile(r'^([A-Za-z]+)(\d+)$')
_RANGE_RE = re.compile(r'^([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)$')


def _validate_cell_ref(ref: str) -> None:
    """Raise ValueError if ref is not a valid A1 cell reference."""
    if not _CELL_RE.match(ref.strip()):
        raise ValueError(f"invalid cell reference: {ref!r}")


def _validate_range_ref(ref: str) -> None:
    """Raise ValueError if ref is not a valid A1 range reference."""
    if not _RANGE_RE.match(ref.strip()):
        raise ValueError(f"invalid range reference: {ref!r}")


def _resolve_ref(wb, current_ws, ref: str):
    """
    Resolve a (possibly sheet-qualified) cell or range ref.
    Returns (worksheet, bare_ref) tuple.
    Raises ValueError for bad refs or missing sheets.
    """
    ref = ref.strip()
    if '!' in ref:
        sheet_name, bare_ref = ref.split('!', 1)
        sheet_name = sheet_name.strip()
        bare_ref = bare_ref.strip()
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"sheet not found: {sheet_name!r} (referenced in {ref!r})")
        ws = wb[sheet_name]
    else:
        ws = current_ws
        bare_ref = ref
    return ws, bare_ref


def _get_cell_value(wb, current_ws, ref: str):
    """
    Read a single cell value from a (possibly sheet-qualified) ref.
    Raises ValueError on bad ref or non-numeric value.
    """
    ws, bare_ref = _resolve_ref(wb, current_ws, ref)
    _validate_cell_ref(bare_ref)
    val = ws[bare_ref].value
    if val is None:
        raise ValueError(f"cell {ref!r} is empty (expected a numeric value)")
    if not isinstance(val, (int, float)):
        raise ValueError(f"cell {ref!r} has non-numeric value: {val!r}")
    return val


def _eval_sum(wb, current_ws, range_ref: str) -> float:
    """
    Sum all numeric cells in range_ref (sheet-qualified or bare).
    Skips blank/None cells. Raises on non-numeric text or empty range.
    """
    ws, bare_range = _resolve_ref(wb, current_ws, range_ref)
    _validate_range_ref(bare_range)

    # openpyxl A1 range slicing returns a tuple of tuples of cells
    cell_range = ws[bare_range]
    # Flatten: ws["B2:B9"] returns tuple of row-tuples
    cells = []
    for row in cell_range:
        if hasattr(row, '__iter__') and not hasattr(row, 'value'):
            cells.extend(row)
        else:
            cells.append(row)

    total = 0.0
    found_any = False
    for cell in cells:
        val = cell.value
        if val is None:
            continue  # skip blank
        if not isinstance(val, (int, float)):
            raise ValueError(
                f"non-numeric value in sum range {range_ref!r}: "
                f"cell {cell.coordinate!r} has value {val!r} "
                f"(check that the range does not include a header row)"
            )
        total += val
        found_any = True

    if not found_any:
        raise ValueError(f"sum range {range_ref!r} is empty or all cells are blank")

    return total


def _eval_computed(wb, current_ws, entry: dict) -> float:
    """Evaluate a single computed cell entry and return the numeric result."""
    op = entry.get("op")
    decimals = entry.get("decimals", 2)

    if op == "sum":
        range_ref = entry.get("range")
        if not range_ref:
            raise ValueError("'sum' op requires a 'range' field")
        result = _eval_sum(wb, current_ws, range_ref)
        return round(result, decimals)

    elif op == "percent_of_total":
        value_ref = entry.get("value_ref")
        total_ref = entry.get("total_ref")
        if not value_ref or not total_ref:
            raise ValueError("'percent_of_total' op requires 'value_ref' and 'total_ref' fields")
        value = _get_cell_value(wb, current_ws, value_ref)
        total = _get_cell_value(wb, current_ws, total_ref)
        if total == 0:
            raise ValueError(
                f"divide-by-zero: total_ref {total_ref!r} is 0 "
                f"(value_ref={value_ref!r})"
            )
        ratio = value / total
        as_fraction = entry.get("as_fraction", False)
        if as_fraction:
            return round(ratio, decimals)
        else:
            return round(ratio * 100, decimals)

    elif op == "ratio":
        value_ref = entry.get("value_ref")
        denominator_ref = entry.get("denominator_ref")
        if not value_ref or not denominator_ref:
            raise ValueError("'ratio' op requires 'value_ref' and 'denominator_ref' fields")
        value = _get_cell_value(wb, current_ws, value_ref)
        denominator = _get_cell_value(wb, current_ws, denominator_ref)
        if denominator == 0:
            raise ValueError(
                f"divide-by-zero: denominator_ref {denominator_ref!r} is 0 "
                f"(value_ref={value_ref!r})"
            )
        return round(value / denominator, decimals)

    else:
        raise ValueError(f"unknown op: {op!r} (expected 'sum', 'percent_of_total', or 'ratio')")


def build(spec: dict) -> None:
    output_path = spec.get("output")
    if not output_path:
        print("Missing 'output' key in input spec", file=sys.stderr)
        sys.exit(1)

    sheets_spec = spec.get("sheets", [])
    if not sheets_spec:
        print("No sheets specified", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    # Remove the default empty sheet so we control creation order.
    wb.remove(wb.active)  # type: ignore[arg-type]

    # ── Pass 1: create every sheet and write rows + cells ────────────────────
    # All sheets and their static values must exist before computed refs resolve.
    for sheet_spec in sheets_spec:
        name = sheet_spec.get("name", "Sheet")
        ws = wb.create_sheet(title=name)

        # Write row-based data first.
        for row_data in sheet_spec.get("rows") or []:
            ws.append(list(row_data))

        # Apply fine-grained cell overrides. Assigning via the A1-style ref
        # (ws["B2"] = ...) avoids depending on openpyxl.utils helpers whose
        # import paths have shifted across openpyxl versions.
        for cell_spec in sheet_spec.get("cells") or []:
            ref = cell_spec.get("ref")
            if not ref:
                continue
            if "formula" in cell_spec:
                ws[ref] = cell_spec["formula"]
            elif "value" in cell_spec:
                ws[ref] = cell_spec["value"]

    # ── Pass 2: evaluate computed cells sequentially ─────────────────────────
    # Walk sheets in creation order. Each computed entry is evaluated and written
    # immediately so that earlier results feed later computations (e.g. a sum
    # written to B10 can be consumed by a subsequent percent_of_total).
    for sheet_spec in sheets_spec:
        name = sheet_spec.get("name", "Sheet")
        ws = wb[name]
        for entry in sheet_spec.get("computed") or []:
            dest_ref = entry.get("ref")
            if not dest_ref:
                print("computed entry missing 'ref' field — skipping", file=sys.stderr)
                sys.exit(1)
            dest_ws, bare_dest = _resolve_ref(wb, ws, dest_ref)
            _validate_cell_ref(bare_dest)
            result = _eval_computed(wb, ws, entry)
            dest_ws[bare_dest] = result

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: build_workbook.py <input.json>", file=sys.stderr)
        sys.exit(1)
    input_path = sys.argv[1]
    with open(input_path, "r", encoding="utf-8") as f:
        spec = json.load(f)
    try:
        build(spec)
    except (ValueError, KeyError) as e:
        print(str(e), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
